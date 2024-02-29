from tkinter import *
from openpyxl import Workbook
from datetime import datetime, timedelta

def set_screen():
    # Set screen icon
    iconphoto = PhotoImage(file = 'twenty-salon\스무살롱.png').subsample(2,2)
    root.wm_iconphoto(False, iconphoto)

    # Set screen title
    root.title("스무살롱")

    # Get screen dimensions
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Setting screen size
    app_width = 280
    app_height = 320

    # Calculate window position
    x = (screen_width - app_width) // 2
    y = (screen_height - app_height) // 2

    # Set screen geometry
    root.geometry(f"{app_width}x{app_height}+{x}+{y}")

def add_name():
    name = entry_name.get().strip() # 이름 입력 받음
    if name:
        list_name.insert(END, name) # 이름 리스트에 추가
        entry_name.delete(0, END)   # 이름 입력 텍스트 상자 초기화
    else:
        entry_name.insert(0, "이름") # 이름이 비어있을 경우 기본값으로 설정

def delete_name():
    list_name.delete(list_name.curselection()) # 선택된 리스트를 삭제

def clear_name():
    list_name.delete(0, END) # Listbox 에 있는 멤버 명단 초기화
    

def add_schedule(year, month):
    # Create a new Workbook
    wb = Workbook()
    ws = wb.active

    # Set properties of WorkSheet
    ws.title = "스무살롱 녹음일정"
    ws.sheet_properties.tabColor = "0888D0" # RGB 형태로 값을 넣어주면 탭 색상 변경

    # Define weekdays
    weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    
    # Set column headers with weekdays
    for i, day in enumerate(weekdays):
        ws.cell(row=1, column=i+1, value=day)
    
    # Get the first day of the month
    first_day = datetime(year, month, 1)

    # Calculate the starting column for the first day
    start_column = first_day.weekday() + 1
    print(first_day.weekday())
    # # 가용 가능한 인원을 기록하는 WorkSheet
    # ws_available = wb.create_sheet("3월 녹음 가용 인원")
    # member = list_name.get(0, END) # 가용 가능 인원
    # ws_available.append(member)

    wb.save("스무살롱_녹음일정.xlsx")
    wb.close()

# Create root window
root = Tk()

# Configure screen setting
set_screen()

# Set frame to manage to member
frame_member = LabelFrame(root, text="녹음 인원 관리")
frame_member.grid(row=0, column=0, rowspan=2, sticky=N+E+W+S, padx=5, pady=5)

# 멤버 목록 표시를 위한 listbox
list_name = Listbox(frame_member, width=15)
list_name.grid(row=0, column=0, columnspan=3, sticky=N+E+W+S, padx=3, pady=3)

# 이름 입력을 위한 텍스트 상자
entry_name = Entry(frame_member, width=10)
entry_name.insert(0, "이름")
entry_name.grid(row=1, column=0, sticky=N+E+W+S, padx=3, pady=3)

# 이름 추가 버튼
button_add = Button(frame_member, text="추가", command=add_name)
button_add.grid(row=1, column=1, padx=3, pady=3)

# 이름 삭제 버튼
button_delete = Button(frame_member, text="삭제", command=delete_name)
button_delete.grid(row=1, column=2, padx=3, pady=3)

# 이름 초기화 버튼
button_clear = Button(frame_member, text="초기화", command=clear_name)
button_clear.grid(row=2, column=0, columnspan=3, sticky=N+E+W+S, padx=3, pady=3)

# # 일정을 추가하려는 시작 회차 입력을 위한 텍스트 상자
# entry_episode = Entry(root, width=10)
# entry_episode.insert(0, "시작 회차")
# entry_episode.grid(row=)

# Set frame to manage number of member
frame_number_member = LabelFrame(root, text="녹음 인원 설정")
frame_number_member.grid(row=0, column=1, padx=5, pady=5)

# Radiobutton row 값 설정
row = 0

# 녹음 인원 수를 2 ~ 6명으로 설정 가능한 Radiobutton 생성
member_var = IntVar()

for i in range(2, 7):
    globals()[f'radio_{i}'] = Radiobutton(frame_number_member, text=f"{i}인", value=i, variable=member_var)
    globals()[f'radio_{i}'].grid(row=row)
    row += 1

    if i == 4: 
        globals()[f'radio_{i}'].select() # default ; 4인

# Set frame to manage number of team
frame_number_team = LabelFrame(root, text="녹음 팀 설정")
frame_number_team.grid(row=1, column=1, sticky=N+E+W+S, padx=5, pady=5)

# 녹음 팀 수를 1 ~ 4팀으로 설정 가능한 Radiobutton 생성
team_var = IntVar()

for i in range(1, 5):
    globals()[f'radio_{i}'] = Radiobutton(frame_number_team, text=f"{i}팀", value=i, variable=team_var)
    globals()[f'radio_{i}'].grid(row=row)
    row += 1

    if i == 2: 
        globals()[f'radio_{i}'].select() # default ; 2팀

# 녹음일정 엑셀 파일에 추가하기
button_add_schedule = Button(root, text="일정 추가", command=add_schedule)
button_add_schedule.grid(row=2, column=0, sticky=N+E+W+S, padx=3, pady=3)

# 앱 종료
button_quit = Button(root, text="종료", command=root.quit)
button_quit.grid(row=2, column=1, sticky=N+E+W+S, padx=3, pady=3)

root.mainloop()